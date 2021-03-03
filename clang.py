from selenium import webdriver
from time import sleep
from selenium.webdriver.support.select import Select #下拉框的选择操作
from selenium.webdriver.common.action_chains import ActionChains
import xlrd2 #xlrd只支持xls，xlrd2可以支持xlsx格式
import os
from openpyxl import load_workbook
import re
from selenium.webdriver.common.keys import Keys


#写入excel的方法
class Write_excel(object):
    '''修改excel数据'''
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active  # 激活sheet

    def write(self, row_n, col_n, value):
        '''写入数据，如(2,3,"hello"),第二行第三列写入数据"hello"'''
        self.ws.cell(row_n, col_n,value )
        self.wb.save(self.filename)

#创建chrome对象，在电脑上打开一个窗口
browser = webdriver.Chrome()
browser.get("https://woger.login.myclang.com/clang/build/production/ClangUI/index.html")
browser.set_window_size(1400,800)
browser.implicitly_wait(50)
browser.find_element_by_id("textfield-1015-inputEl").send_keys("mara.shu@HabaTrading.com")
browser.find_element_by_id("textfield-1016-inputEl").send_keys("L2cLE44Ry1VX")
browser.execute_script("arguments[0].click();", browser.find_element_by_id("button-1020-btnInnerEl"))
browser.implicitly_wait(300)


def tag_me():
    browser.find_element_by_xpath("//table[@data-recordindex='10']").click()
    tag = browser.find_element_by_xpath("//input[@name='tags']")
    tag.send_keys("Home")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Banks")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='11']"))
    tag.send_keys("Home & garden")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Home & garden accessories")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("rugs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='12']"))
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Chairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("dining room chairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='13']"))
    tag.send_keys("Home & garden")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Lighting")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("lamps")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='14']"))
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("chairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("folding seats & stools")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='15']"))
    tag.send_keys("Home & garden")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Home & garden accessories")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Decorative pillows")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='16']"))
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Cabinets & storage furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Buffets & sideboards")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='17']"))
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("chairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("dining room chairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='18']"))
    tag.send_keys("Home & garden")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Home & garden accessories")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("rugs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='19']"))
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("chairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Armchairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='20']"))
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("chairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Armchairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='21']"))
    tag.send_keys("Home & garden")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Lighting")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("lamps")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='29']"))
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("tables")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("bedside tables")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='30']"))
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Cabinets & storage furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Buffets & sideboards")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='31']"))
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Beds & accessories")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Beds & bed frames")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='32']"))
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Cabinets & storage furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Buffets & sideboards")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='33']"))
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Chairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Armchairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='34']"))
    tag.send_keys("Home & garden")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Linen")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Bed linen")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)

    #最后一排
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='36']"))
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("tables")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Dining tables")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='37']"))
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("tables")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Dining tables")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='38']"))
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("tables")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Dining tables")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='39']"))
    tag.send_keys("Home & garden")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Home & garden accessories")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Mirrors")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='40']"))
    tag.send_keys("Home & garden")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Home & garden accessories")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Mirrors")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='41']"))
    tag.send_keys("Home & garden")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Home & garden accessories")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Mirrors")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='42']"))
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("chairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("dining room chairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='43']"))
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("chairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("dining room chairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='44']"))
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("chairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("dining room chairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='45']"))
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("chairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Armchairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='46']"))
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("chairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Armchairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//table[@data-recordindex='47']"))
    tag.send_keys("Furniture")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("chairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)
    tag.send_keys("Armchairs")  # 添加tag然后回车
    tag.send_keys(Keys.ENTER)

#获取表格数据
def obtain_table_data(url):
    # 获取表格数据.打开文件
    current_path = os.path.dirname(__file__)
    excel_paht = os.path.join(current_path,url) #r是转义保持原文件路径
    workbook = xlrd2.open_workbook(excel_paht)
    sheet=workbook.sheet_by_index(0)
    #双列表形式 ，一行一个用例
    all_case_info = []
    for i in range(1, sheet.nrows):  # 行数所以从1开始，0一般是名称之类的
        case_info = []
        for j in range(sheet.ncols):
            case_info.append(sheet.cell_value(i, j))
        all_case_info.append(case_info)  # 注意，python以对齐来确定循环的所定义区域
    return all_case_info

data_table = obtain_table_data("C:\\Users\\Mara.Shu\\Desktop\\clang_test.xlsx")



def main():
    browser.find_element_by_xpath("//li[@data-recordid='152']").click()
    browser.find_element_by_xpath("//input[@data-ref='inputEl']").send_keys("newsletter 2021")
    ActionChains(browser).double_click(browser.find_element_by_xpath("//div[@data-attr-file-id='67297']")).perform() #鼠标双击操作
    ActionChains(browser).double_click(browser.find_element_by_xpath("//div[@data-attr-file-id='68595']")).perform() #
    ActionChains(browser).double_click(browser.find_element_by_xpath("//div[@data-attr-file-id='68596']")).perform() #
    ActionChains(browser).double_click(browser.find_element_by_xpath("//div[@data-qtip='Newsletter week 09 follow the trend AT']")).perform() #
    browser.execute_script("arguments[0].click();", browser.find_element_by_link_text("Clickthroughs"))  #点击clickthrough
    browser.implicitly_wait(50)
    browser.execute_script("arguments[0].click();", browser.find_element_by_link_text("Refresh"))  #点击刷新
    browser.execute_script("arguments[0].click();", browser.find_element_by_link_text("Continue"))  #跳过刷新
    browser.execute_script("arguments[0].click();", browser.find_element_by_css_selector("tr.x-grid-row:first-child"))
    browser.execute_script("arguments[0].click();", browser.find_element_by_link_text("Select all")) #点击select all
    browser.implicitly_wait(50)
    browser.execute_script("arguments[0].click();", browser.find_element_by_xpath("//input[@name='conversion_tracking']"))
    browser.find_element_by_xpath("//input[@name='google_source']").send_keys("vidaxl_newsletter")
    browser.find_element_by_xpath("//input[@name='google_medium']").send_keys("email")
    browser.find_element_by_xpath("//input[@name='google_campaign']").send_keys("Newsletter week 09 follow the trend AT 2021")
    browser.find_element_by_xpath("//input[@name='tags']").send_keys("Inspiration") #添加tag然后回车
    browser.find_element_by_xpath("//input[@name='tags']").send_keys(Keys.ENTER)
    browser.implicitly_wait(50)
    tag_me()
    browser.implicitly_wait(50)
    # browser.execute_script("arguments[0].click();", browser.find_element_by_link_text("Save"))  # 点击保存邮件

main()
we = Write_excel("C:\\Users\\Mara.Shu\\Desktop\\clang_test.xlsx")
we.write(2, 5, "1")
browser.implicitly_wait(50)
