from selenium import webdriver
from time import sleep
from selenium.webdriver.support.select import Select #下拉框的选择操作
from selenium.webdriver.common.action_chains import ActionChains
import xlrd2 #xlrd只支持xls，xlrd2可以支持xlsx格式
import os
from openpyxl import load_workbook
import re

# 注意事项
# 1、文件的country那一列不能更改，顺序无所谓，第一排名字可以改，但是要看得懂
# 2、图片的那一列需要从\换成\\，不然上传会因为有空格等因素出错
# 3、description这列有数据就放，没有需要删掉，然后代码中加成空格。
# 4、添加的时候日期需要确认好。上传需要修改文件的名字


#写入excel的方法
class Write_excel(object):
    '''修改excel数据'''
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active  # 激活sheet

    def write(self, row_n, col_n, value):
        '''写入数据，如(2,3,"hello"),第二行第三列写入数据"hello"'''
        self.ws.cell(row_n, col_n,value )
        self.wb.save(self.filename)

#创建chrome对象，在电脑上打开一个窗口
browser = webdriver.Chrome()
browser.get("https://admin.woger-cdn.com/index.php/admin/dashboard/index/key/ae57004afb8c09ab3969ca72b85962f7/")
sleep(2)
browser.set_window_size(1400,800)
browser.find_element_by_id("username").send_keys("Gloria.Wu@habatrading.com")
browser.find_element_by_id("login").send_keys("!gWJP!K8j*3%LG")
browser.find_element_by_class_name("form-button").click()
browser.implicitly_wait(50)
browser.refresh()
#定位到要悬停的元素
browser.implicitly_wait(50)
link = browser.find_element_by_link_text("Sales")
ActionChains(browser).move_to_element(link).perform()
Landingpage = browser.find_element_by_link_text("Landingpage")
ActionChains(browser).move_to_element(Landingpage).perform()
browser.implicitly_wait(50) #属于隐式等待，5秒钟内只要找到了元素就开始执行，5秒钟后未找到，就超时
#点开需要悬停之后出现的元素
browser.find_element_by_link_text("Landing Pages").click()

#添加landing page的方法动作和传值
def by_value(storeid,title,link,stock,top_image_file,description):
    browser.implicitly_wait(60)
    browser.find_element_by_xpath('//*[@id="page:main-container"]/div[2]/table/tbody/tr/td[2]/button[@title="Add Landingpage"]').click()  # 选择class是scalable的第一个参数
    #通过id然后找到option的value只然后选定国家，这是下拉框的选项select_by_value(“选择值”)select_by_index(“索引值”)select_by_visible_text(“文本值”)
    store_id = browser.find_element_by_id("store_id")
    Select(store_id).select_by_visible_text(storeid)
    browser.find_element_by_id("title").send_keys(title) # 根据表格的数据像title传值
    browser.find_element_by_id("url").send_keys(link)
    browser.find_element_by_id("weight").send_keys("10")
    browser.find_element_by_id("date_from").send_keys("2021-03-23 01:28:33")
    browser.find_element_by_id("date_to").send_keys("2031-12-31 07:28:33")
    #定位上传文件按钮，添加本地文件
    browser.find_element_by_id("top_image_file").send_keys(top_image_file)
    browser.find_element_by_id("query").send_keys(stock)
    status = browser.find_element_by_id("status")  # 选择状态
    Select(status).select_by_value("1")
    browser.find_element_by_id("display_category_tree").click() #选择是否category tree

    browser.implicitly_wait(50)
    browser.switch_to.frame("description_ifr")
    browser.find_element_by_css_selector("#tinymce p").send_keys(description)
    browser.switch_to.default_content()
    browser.find_element_by_class_name("save").click()
    sleep(2)

#传id的方法动作和传值
def by_id(storeid,link):
    browser.find_element_by_id("landingpageId_filter_url").clear() #先清除已经加过的文字，在输入文字
    browser.find_element_by_id("landingpageId_filter_url").send_keys(link)
    browser.find_element_by_css_selector("button.task").click()
    # 根据文本定位点击的国家
    tr = browser.find_elements_by_css_selector("table#landingpageId_table tbody tr")
    for len_tr in range(0, len(tr)):
        text = tr[len_tr].get_attribute("textContent")  # 获取元素标签的内容textContent 获取元素内的全部HTML innerHTML 获取包含选中元素的HTML outerHTML
        if text.find(storeid) != -1:
            tr[len_tr].click()
            break

#获取表格数据
def obtain_table_data(url):
    # 获取表格数据.打开文件
    current_path = os.path.dirname(__file__)
    excel_paht = os.path.join(current_path,url) #r是转义保持原文件路径
    workbook = xlrd2.open_workbook(excel_paht)
    sheet=workbook.sheet_by_index(0)
    #双列表形式 ，一行一个用例
    all_case_info = []
    for i in range(1, sheet.nrows):  # 行数所以从1开始，0一般是名称之类的
        case_info = []
        for j in range(sheet.ncols):
            case_info.append(sheet.cell_value(i, j))
        all_case_info.append(case_info)  # 注意，python以对齐来确定循环的所定义区域
    return all_case_info

data_table = obtain_table_data("C:\\Users\\Mara.Shu\\Desktop\\7. Home office.xlsx")

#传入文件数据，添加landing page
for num in range(0,len(data_table)):
    by_value(data_table[num][1],data_table[num][2],data_table[num][3],data_table[num][4],data_table[num][5],data_table[num][6])

browser.implicitly_wait(50)

#传输id到excel里面
# rule = r'id/(.*?)/'
# for num in range(0,len(data_table)):
#     by_id(data_table[num][1],data_table[num][3])
#     id_data = re.findall(rule, browser.current_url) #用正则的方法取出其中的id，正则表达式之后的数据显示的是一个数组，所以需要提取第一个数据
#     we = Write_excel("C:\\Users\\Mara.Shu\\Desktop\\4. Head - Rattan outdoor funriture.xlsx")
#     we.write(num+2,8,id_data[0])
#     browser.find_element_by_css_selector("button.back:first-child").click()
#     browser.implicitly_wait(50)